import io
from datetime import date, timedelta
from flask import Blueprint, request, jsonify, send_file
from sqlalchemy import func
from extensions import db
from models import Submission, Participant, ReviewLog, Challenge
from utils.dates import get_today_local

bp = Blueprint("admin", __name__, url_prefix="/api/admin")


# ─── Review ──────────────────────────────────────────────────────────────────

@bp.route("/submission/<int:sid>/review", methods=["PUT"])
def review_submission(sid):
    submission = Submission.query.get_or_404(sid)
    data = request.get_json()

    # Capture old state for audit
    old = {
        "review_status": submission.review_status,
        "approved_steps": submission.approved_steps,
        "fitness_video_bonus": submission.fitness_video_bonus,
        "fitness_attendance_bonus": submission.fitness_attendance_bonus,
        "food_bonus": submission.food_bonus,
        "total_points": submission.total_points,
    }

    submission.review_status = data.get("review_status", "Approved")
    submission.approved_steps = int(data.get("approved_steps", submission.steps))
    submission.fitness_video_bonus = int(data.get("fitness_video_bonus", 0))
    submission.fitness_attendance_bonus = int(data.get("fitness_attendance_bonus", 0))
    submission.food_bonus = int(data.get("food_bonus", 0))
    submission.reviewed_by = data.get("reviewer", "Admin")
    submission.recalculate_points()

    new = {
        "review_status": submission.review_status,
        "approved_steps": submission.approved_steps,
        "fitness_video_bonus": submission.fitness_video_bonus,
        "fitness_attendance_bonus": submission.fitness_attendance_bonus,
        "food_bonus": submission.food_bonus,
        "total_points": submission.total_points,
    }

    log = ReviewLog(
        submission_id=sid,
        reviewer=data.get("reviewer", "Admin"),
        action=submission.review_status,
    )
    log.set_old_values(old)
    log.set_new_values(new)
    db.session.add(log)
    db.session.commit()

    return jsonify(submission.to_dict())


@bp.route("/submission/<int:sid>/audit", methods=["GET"])
def audit_log(sid):
    logs = ReviewLog.query.filter_by(submission_id=sid).order_by(ReviewLog.timestamp.desc()).all()
    return jsonify([l.to_dict() for l in logs])


# ─── Stats ────────────────────────────────────────────────────────────────────

@bp.route("/stats", methods=["GET"])
def dashboard_stats():
    challenge_id = request.args.get("challenge_id")
    base = Submission.query
    if challenge_id:
        base = base.filter_by(challenge_id=challenge_id)

    total = base.count()
    pending = base.filter(Submission.review_status == "Pending").count()
    approved = base.filter(Submission.review_status == "Approved").count()
    rejected = base.filter(Submission.review_status == "Rejected").count()

    avg_pts = (
        db.session.query(func.avg(Submission.total_points))
        .filter(Submission.review_status == "Approved")
        .scalar()
    )
    max_pts = (
        db.session.query(func.max(Submission.total_points))
        .filter(Submission.review_status == "Approved")
        .scalar()
    )

    return jsonify({
        "total": total,
        "pending": pending,
        "approved": approved,
        "rejected": rejected,
        "avg_points": round(avg_pts or 0, 1),
        "max_points": max_pts or 0,
    })


# ─── Analytics ───────────────────────────────────────────────────────────────

@bp.route("/analytics", methods=["GET"])
def analytics():
    challenge_id = request.args.get("challenge_id")

    # Top performers
    top_q = (
        db.session.query(
            Participant.name,
            func.sum(Submission.total_points).label("pts"),
        )
        .join(Submission, Submission.participant_id == Participant.id)
        .filter(Submission.review_status == "Approved")
    )
    if challenge_id:
        top_q = top_q.filter(Submission.challenge_id == challenge_id)
    top_performers = [
        {"name": r.name, "points": int(r.pts)}
        for r in top_q.group_by(Participant.id).order_by(func.sum(Submission.total_points).desc()).limit(10)
    ]

    # Submission trend (last 14 days)
    trend = []
    for i in range(13, -1, -1):
        day = get_today_local() - timedelta(days=i)
        count = Submission.query.filter(Submission.submission_date == day).count()
        trend.append({"date": day.isoformat(), "count": count})

    # Approval rate
    total = Submission.query.count()
    approved = Submission.query.filter_by(review_status="Approved").count()
    rejected = Submission.query.filter_by(review_status="Rejected").count()
    pending = Submission.query.filter_by(review_status="Pending").count()

    # Points distribution buckets
    buckets = {"0-5000": 0, "5001-10000": 0, "10001-15000": 0, "15001+": 0}
    pts_rows = Submission.query.filter_by(review_status="Approved").with_entities(Submission.total_points).all()
    for (pts,) in pts_rows:
        if pts <= 5000:
            buckets["0-5000"] += 1
        elif pts <= 10000:
            buckets["5001-10000"] += 1
        elif pts <= 15000:
            buckets["10001-15000"] += 1
        else:
            buckets["15001+"] += 1

    return jsonify({
        "top_performers": top_performers,
        "submission_trend": trend,
        "approval_rate": {
            "total": total,
            "approved": approved,
            "rejected": rejected,
            "pending": pending,
        },
        "points_distribution": [{"range": k, "count": v} for k, v in buckets.items()],
    })


# ─── Excel Export ─────────────────────────────────────────────────────────────

@bp.route("/export/<int:challenge_id>", methods=["GET"])
def export_excel(challenge_id):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"message": "openpyxl not installed"}), 500

    challenge = Challenge.query.get_or_404(challenge_id)

    # Get day names for the challenge window
    delta = (challenge.end_date - challenge.start_date).days + 1
    day_dates = [challenge.start_date + timedelta(days=i) for i in range(delta)]

    # Aggregate points per participant per day
    submissions = (
        Submission.query
        .filter_by(challenge_id=challenge_id, review_status="Approved")
        .all()
    )

    data: dict[str, dict] = {}
    for s in submissions:
        eid = s.participant.employee_id
        if eid not in data:
            data[eid] = {"name": s.participant.name, "days": {}}
        data[eid]["days"][s.submission_date] = s.total_points

    # Build leaderboard
    totals = {eid: sum(v["days"].values()) for eid, v in data.items()}
    ranked = sorted(totals.items(), key=lambda x: x[1], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = challenge.name[:31]

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    # Header row
    headers = ["Rank", "Employee ID", "Name"] + [d.strftime("%a %d %b") for d in day_dates] + ["Weekly Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, (eid, total) in enumerate(ranked, 2):
        info = data[eid]
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=eid)
        ws.cell(row=row_idx, column=3, value=info["name"])
        for col_idx, d in enumerate(day_dates, 4):
            ws.cell(row=row_idx, column=col_idx, value=info["days"].get(d, 0))
        ws.cell(row=row_idx, column=4 + len(day_dates), value=total)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ShapeUp_{challenge.name.replace(' ', '_')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)